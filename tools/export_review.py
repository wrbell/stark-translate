"""Export batch translation results to a formatted Excel workbook for bilingual review.

Reads segments.jsonl files produced by batch_translate.py and generates a single
.xlsx workbook with one sheet per sermon plus a summary sheet. Columns show source
text alongside MarianMT and Gemma 4B translations for side-by-side comparison.

Usage:
    python tools/export_review.py stark_data/arlington_2026/
    python tools/export_review.py stark_data/arlington_2026/ --output review.xlsx
    python tools/export_review.py stark_data/arlington_2026/february_14_2026_1_.../
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -- Styles ------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

_COLUMNS = [
    ("#", 6),
    ("Time", 12),
    ("Duration", 9),
    ("English (STT)", 55),
    ("Spanish (MarianMT)", 55),
    ("Spanish (Gemma 4B)", 55),
    ("STT Conf", 9),
    ("Notes", 30),
]


# -- Helpers ------------------------------------------------------------------


def _fmt_time(seconds: float) -> str:
    """Format seconds as MM:SS."""
    m, s = divmod(int(seconds), 60)
    return f"{m:02d}:{s:02d}"


def _load_segments(jsonl_path: Path) -> list[dict]:
    """Load and sort segment records from a JSONL file."""
    records = []
    with open(jsonl_path) as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    records.sort(key=lambda r: r.get("segment_index", 0))
    return records


def _truncate_sheet_name(name: str, max_len: int = 31) -> str:
    """Truncate to Excel's 31-character sheet name limit."""
    if len(name) <= max_len:
        return name
    return name[: max_len - 1] + "\u2026"


def _find_sermon_dirs(input_path: Path) -> list[Path]:
    """Find directories containing segments.jsonl files."""
    # Single sermon directory
    if (input_path / "segments.jsonl").exists():
        return [input_path]
    # Batch directory — find all subdirs with segments.jsonl
    dirs = sorted(d for d in input_path.iterdir() if d.is_dir() and (d / "segments.jsonl").exists())
    return dirs


def _apply_header(ws, col_count: int) -> None:
    """Style the header row and set column widths."""
    for col_idx, (header, width) in enumerate(_COLUMNS[:col_count], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _confidence_fill(conf: float | None) -> PatternFill | None:
    """Return fill color based on STT confidence value."""
    if conf is None:
        return None
    if conf >= 0.9:
        return _GREEN_FILL
    if conf >= 0.7:
        return _YELLOW_FILL
    return _RED_FILL


# -- Sheet builders -----------------------------------------------------------


def _build_sermon_sheet(ws, segments: list[dict]) -> dict:
    """Populate a worksheet with segment data. Returns summary stats."""
    col_count = len(_COLUMNS)
    _apply_header(ws, col_count)

    total_duration = 0.0
    conf_sum = 0.0
    conf_count = 0

    for row_idx, seg in enumerate(segments, start=2):
        start = seg.get("start_sec", 0.0)
        end = seg.get("end_sec", 0.0)
        dur = seg.get("duration_sec", 0.0)
        conf = seg.get("stt_confidence")
        total_duration += dur

        if conf is not None:
            conf_sum += conf
            conf_count += 1

        values = [
            seg.get("segment_index", row_idx - 2),
            f"{_fmt_time(start)}-{_fmt_time(end)}",
            f"{dur:.1f}s",
            seg.get("source_text", ""),
            seg.get("marian_text", ""),
            seg.get("gemma_text", ""),
            f"{conf:.2f}" if conf is not None else "",
            "",  # Notes column — empty for reviewer
        ]

        is_alt_row = (row_idx % 2) == 0
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _WRAP
            if is_alt_row:
                cell.fill = _ALT_ROW_FILL

        # Confidence conditional coloring (overrides alt-row shading)
        conf_cell = ws.cell(row=row_idx, column=7)
        fill = _confidence_fill(conf)
        if fill:
            conf_cell.fill = fill

    return {
        "segments": len(segments),
        "duration": total_duration,
        "avg_confidence": (conf_sum / conf_count) if conf_count > 0 else None,
    }


def _build_summary_sheet(ws, sermon_stats: list[tuple[str, dict]]) -> None:
    """Build the summary sheet with per-sermon stats."""
    headers = [
        ("Sermon", 50),
        ("Segments", 10),
        ("Duration", 12),
        ("Avg Confidence", 14),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    total_segs = 0
    total_dur = 0.0

    for row_idx, (name, stats) in enumerate(sermon_stats, start=2):
        dur = stats["duration"]
        avg_conf = stats["avg_confidence"]
        total_segs += stats["segments"]
        total_dur += dur

        dur_min = int(dur) // 60
        dur_sec = int(dur) % 60

        values = [
            name.replace("_", " ").title(),
            stats["segments"],
            f"{dur_min}m {dur_sec}s",
            f"{avg_conf:.2f}" if avg_conf is not None else "N/A",
        ]
        is_alt = (row_idx % 2) == 0
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _WRAP
            if is_alt:
                cell.fill = _ALT_ROW_FILL

    # Totals row
    totals_row = len(sermon_stats) + 2
    total_min = int(total_dur) // 60
    total_sec = int(total_dur) % 60
    totals = ["TOTAL", total_segs, f"{total_min}m {total_sec}s", ""]
    for col_idx, val in enumerate(totals, start=1):
        cell = ws.cell(row=totals_row, column=col_idx, value=val)
        cell.font = Font(bold=True, size=11)
        cell.alignment = _WRAP


# -- Main ---------------------------------------------------------------------


def export_review(input_path: Path, output_path: Path | None = None) -> Path:
    """Generate a review Excel workbook from batch translation output.

    Args:
        input_path: Batch output directory or single sermon subdirectory.
        output_path: Output .xlsx path. Defaults to <input_path>/review.xlsx.

    Returns:
        Path to the generated .xlsx file.
    """
    sermon_dirs = _find_sermon_dirs(input_path)
    if not sermon_dirs:
        print(f"No segments.jsonl files found in {input_path}", file=sys.stderr)
        sys.exit(1)

    if output_path is None:
        output_path = input_path / "review.xlsx"

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    wb.remove(wb.active)

    # Summary sheet first
    summary_ws = wb.create_sheet(title="Summary")
    sermon_stats: list[tuple[str, dict]] = []

    for sermon_dir in sermon_dirs:
        jsonl_path = sermon_dir / "segments.jsonl"
        segments = _load_segments(jsonl_path)
        if not segments:
            continue

        sheet_name = _truncate_sheet_name(sermon_dir.name)
        # Deduplicate sheet names if needed
        existing = {ws.title for ws in wb.worksheets}
        if sheet_name in existing:
            sheet_name = _truncate_sheet_name(sermon_dir.name, max_len=28) + f"_{len(existing)}"

        ws = wb.create_sheet(title=sheet_name)
        stats = _build_sermon_sheet(ws, segments)
        sermon_stats.append((sermon_dir.name, stats))

    _build_summary_sheet(summary_ws, sermon_stats)

    # Move summary to first position
    wb.move_sheet(summary_ws, offset=-(len(wb.worksheets) - 1))

    wb.save(str(output_path))
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export batch translation results to Excel for bilingual review.",
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="Batch output directory or single sermon subdirectory",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: <input_path>/review.xlsx)",
    )
    args = parser.parse_args()

    xlsx_path = export_review(args.input_path, args.output)
    print(f"Review workbook saved to {xlsx_path}")


if __name__ == "__main__":
    main()
