"""Tests for tools/export_review.py — Excel review workbook export."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

pytest.importorskip("openpyxl")

from openpyxl import Workbook

from tools import export_review as er


class TestHelpers:
    def test_fmt_time(self):
        assert er._fmt_time(0) == "00:00"
        assert er._fmt_time(65) == "01:05"
        assert er._fmt_time(601) == "10:01"

    def test_truncate_sheet_name(self):
        assert er._truncate_sheet_name("short") == "short"
        long = "a" * 40
        out = er._truncate_sheet_name(long)
        assert len(out) == 31
        assert out.endswith("\u2026")

    def test_confidence_fill(self):
        assert er._confidence_fill(None) is None
        assert er._confidence_fill(0.95) is er._GREEN_FILL
        assert er._confidence_fill(0.75) is er._YELLOW_FILL
        assert er._confidence_fill(0.4) is er._RED_FILL

    def test_load_segments_sorted(self, tmp_path: Path):
        p = tmp_path / "segments.jsonl"
        rows = [
            {"segment_index": 2, "source_text": "c"},
            {"segment_index": 0, "source_text": "a"},
            {"segment_index": 1, "source_text": "b"},
        ]
        p.write_text("\n".join(json.dumps(r) for r in rows) + "\n\n")
        loaded = er._load_segments(p)
        assert [r["source_text"] for r in loaded] == ["a", "b", "c"]


class TestFindSermonDirs:
    def test_single_dir(self, tmp_path: Path):
        (tmp_path / "segments.jsonl").write_text("{}\n")
        assert er._find_sermon_dirs(tmp_path) == [tmp_path]

    def test_batch_dirs(self, tmp_path: Path):
        a = tmp_path / "sermon_a"
        b = tmp_path / "sermon_b"
        a.mkdir()
        b.mkdir()
        (a / "segments.jsonl").write_text("{}\n")
        (b / "segments.jsonl").write_text("{}\n")
        (tmp_path / "empty").mkdir()
        dirs = er._find_sermon_dirs(tmp_path)
        assert dirs == [a, b]


class TestBuildSheets:
    def test_sermon_and_summary(self):
        wb = Workbook()
        ws = wb.active
        segments = [
            {
                "segment_index": 0,
                "start_sec": 0,
                "end_sec": 5,
                "duration_sec": 5.0,
                "source_text": "Hello",
                "marian_text": "Hola",
                "gemma_text": "Hola",
                "stt_confidence": 0.95,
            },
            {
                "segment_index": 1,
                "start_sec": 5,
                "end_sec": 10,
                "duration_sec": 5.0,
                "source_text": "World",
                "marian_text": "Mundo",
                "gemma_text": "Mundo",
                "stt_confidence": 0.5,
            },
            {
                "segment_index": 2,
                "start_sec": 10,
                "end_sec": 12,
                "duration_sec": 2.0,
                "source_text": "No conf",
                "marian_text": "",
                "gemma_text": "",
                "stt_confidence": None,
            },
        ]
        stats = er._build_sermon_sheet(ws, segments)
        assert stats["segments"] == 3
        assert stats["duration"] == 12.0
        assert stats["avg_confidence"] == pytest.approx((0.95 + 0.5) / 2)

        summary = wb.create_sheet("Summary")
        er._build_summary_sheet(summary, [("sermon_one", stats)])
        assert summary.cell(row=2, column=2).value == 3
        assert summary.cell(row=3, column=1).value == "TOTAL"


class TestExportReview:
    def test_export_workbook(self, tmp_path: Path):
        sermon = tmp_path / "february_14_2026"
        sermon.mkdir()
        segs = [
            {
                "segment_index": 0,
                "start_sec": 0,
                "end_sec": 3,
                "duration_sec": 3.0,
                "source_text": "Grace",
                "marian_text": "Gracia",
                "gemma_text": "Gracia",
                "stt_confidence": 0.9,
            }
        ]
        (sermon / "segments.jsonl").write_text(json.dumps(segs[0]) + "\n")
        out = tmp_path / "review.xlsx"
        path = er.export_review(tmp_path, out)
        assert path == out
        assert out.exists()

    def test_no_segments_exits(self, tmp_path: Path):
        with pytest.raises(SystemExit):
            er.export_review(tmp_path)

    def test_main(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys):
        sermon = tmp_path / "s1"
        sermon.mkdir()
        (sermon / "segments.jsonl").write_text(
            json.dumps(
                {
                    "segment_index": 0,
                    "start_sec": 0,
                    "end_sec": 1,
                    "duration_sec": 1.0,
                    "source_text": "x",
                    "marian_text": "y",
                    "gemma_text": "z",
                    "stt_confidence": 0.8,
                }
            )
            + "\n"
        )
        out = tmp_path / "out.xlsx"
        monkeypatch.setattr(sys, "argv", ["export_review.py", str(tmp_path), "--output", str(out)])
        er.main()
        assert out.exists()
        assert "Review workbook saved" in capsys.readouterr().out
